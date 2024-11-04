import openpyxl as xl
from html2image import Html2Image
import os

def gen_acreditacion(data, f_output):
    hti = Html2Image(output_path=f_output)
    with open('base.html', 'r') as file:
        html = file.read()


    # format the file with the corresponding data
    if "escuela" in data["centro"].lower():
        data["color"] = "#3BC4A0"
        data["centro"] = "eps"
    elif "colme" in data["centro"].lower():
        data["color"] = "#09725F"
        data["centro"] = "colme"
    elif "facultad" in data["centro"].lower():
        data["color"] = "#5599DD"
        data["centro"] = "getafe"
    elif "postgrado" in data["centro"].lower():
        data["color"] = "#FF6D2E"
        data["centro"] = "puerta"
    elif data["centro"] == "SECTORIAL" or data["centro"] == "AUTORIDAD":
        data["color"] = "#1A2A84"

    if data["rol"] == "ORGANIZADOR":
        data["rol"] = """
                        <div style="width: 1031px; height: 176px; left: 380px; top: 710px; position: absolute">
                        <div style="width: 1031px; height: 176px; left: 0px; top: 0px; position: absolute; background: #001489"></div>
                        <div style="width: 835px; height: 94px; left: 83px; top: 48px; position: absolute; text-align: center; color: white; font-size: 70px; font-family: Inter; font-weight: 800; word-wrap: break-word">ORGANIZACIÓN</div>
                        </div>
                        """
    # Ponentes
    elif data["rol"] == "PONENTE":
        data["rol"] = """
                        <div style="width: 1031px; height: 176px; left: 380px; top: 710px; position: absolute">
                        <div style="width: 1031px; height: 176px; left: 0px; top: 0px; position: absolute; background: rgba(0, 20, 137, 0.70)"></div>
                        <div style="width: 835px; height: 94px; left: 83px; top: 48px; position: absolute; text-align: center; color: white; font-size: 70px; font-family: Inter; font-weight: 800; word-wrap: break-word">PONENTE</div>
                        </div>
                        """
    # Sectoriales
    elif data["centro"] == "SECTORIAL":
        data["degree"] = f"PONENTE JFDE <br/> {data['degree']}"

    # FIGURAS DE EQUIPO
    if data["team"] == "AZUL":
        data["team"] = """<div style="width: 111px; height: 111px; right: 30px; top: 30px; position: absolute; background: #2B27D7"></div>"""
    elif data["team"] == "VERDE":
        data["team"] = """<div style="width: 111px; height: 111px; right: 70px; top: 90px; position: absolute; transform: rotate(-44.65deg); transform-origin: 0 0; background: #65CC76"></div>"""
    elif data["team"] == "ROJO":
        data["team"] = """<div style="width: 76px; height: 76px; right: 30px; top: 30px; position: absolute; background: #FF0707; border-radius: 9999px"></div>"""
    elif data["team"] == "AMARILLO":
        data["team"] = """<div style="width: 0; height: 0;right: 30px; top: 30px; border-left: 50px solid transparent; border-right: 50px solid transparent; border-bottom: 100px solid #FFEB3A; position: absolute;"></div>"""
    else:
        data["team"] = "white"

    #data["name"] = data["name"].capitalize()
    html = html.format(**data)
    data['name'] = data['name'].replace(" ", "_")
    data['surname'] = data['surname'].replace(" ", "_")
    if data['surname'][-1] == "_":
        data['surname'] = data['surname'][:-1]
    file_name = f"acred_{data['name']}_{data['surname']}.png"

    # Construye la ruta completa para las imágenes
    ruta_imagen = os.path.join(os.getcwd(), "base", "dele.png")
    ruta_centro_imagen = os.path.join(os.getcwd(), "base", f'{data["centro"]}.png')
    ruta_uc3m_imagen = os.path.join(os.getcwd(), "base", "uc3m.png")
    # Reemplaza las rutas relativas en el HTML con las rutas absolutas
    html = html.replace('src="base/centro.png"', f'src="{ruta_centro_imagen}"')
    html = html.replace('src="base/uc3m.png"', f'src="{ruta_uc3m_imagen}"')
    html = html.replace('src="base/dele.png"', f'src="{ruta_imagen}"')
    print(ruta_imagen)
    print(ruta_centro_imagen)
    print(ruta_uc3m_imagen)

    print(f"Generating {file_name}")
    hti.screenshot(html_str=html, save_as=file_name, size=(1411, 886))


def get_excel_data(excel_path):
    dataframe = xl.load_workbook(excel_path)
    dataframe = dataframe.active
    data = []
    for i in dataframe.iter_rows():
        data.append({dataframe.cell(row=1, column=1).value: i[0].value, dataframe.cell(row=1, column=2).value: i[1].value,
                     dataframe.cell(row=1, column=3).value: i[2].value, dataframe.cell(row=1, column=4).value: i[3].value,
                     dataframe.cell(row=1, column=5).value: i[4].value, dataframe.cell(row=1, column=6).value: i[5].value})
    data.pop(0)
    return data
