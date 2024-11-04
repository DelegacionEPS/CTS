import openpyxl as xl

excel_file_path = 'book.xlsx'
dataframe = xl.load_workbook(excel_file_path)
dataframe = dataframe.active
data = []
for i in dataframe.iter_rows():
    data.append({dataframe.cell(row=1, column=1).value: i[0].value, dataframe.cell(row=1, column=2).value: i[1].value,
               dataframe.cell(row=1, column=3).value: i[2].value, dataframe.cell(row=1, column=4).value: i[3].value,
               dataframe.cell(row=1, column=5).value: i[4].value, dataframe.cell(row=1, column=6).value: i[5].value})
data.pop(0)
